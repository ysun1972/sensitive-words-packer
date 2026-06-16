"""
Excel 词表读取（.xlsx）

支持两种布局：
1. 默认 Sheet 第 1 列：每行一个词（最简）
2. 自定义列名/Sheet：按 header 取指定列

约定：
- .xlsx 文件第一个非空工作表为默认数据源
- 若表头行存在（首行全为字符串），按 header 取列；否则取第 1 列
- 跳过空行、# 注释行
- 返回 list[str]

依赖：openpyxl（PyInstaller 打包需 --collect-all openpyxl）
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable


def _is_comment(cell: str) -> bool:
    s = cell.strip()
    return not s or s.startswith("#")


def _col_letter_to_index(letter: str) -> int:
    """A -> 0, B -> 1, AA -> 26 ..."""
    s = letter.strip().upper()
    n = 0
    for ch in s:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"非法列字母: {letter!r}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _is_header_row(row: Iterable) -> bool:
    """启发式：首行所有 cell 都是 str 类型 → 视为表头"""
    cells = [c for c in row if c is not None and str(c).strip()]
    return len(cells) > 0 and all(isinstance(c, str) for c in cells)


def load_excel_words(
    path: str | Path,
    sheet: str | int | None = None,
    column: str | int | None = None,
) -> list[str]:
    """
    从 .xlsx 文件加载敏感词列表。

    参数:
        path: .xlsx 文件路径
        sheet: 工作表名（str）或索引（int，从 0 开始），None 取第一个非空工作表
        column: 列名（str，匹配 header）或列号（int，从 0 开始），None 取第 1 列
                传 "A"、"B" 等列字母也支持

    返回:
        list[str]，去重 + 去空 + 去注释
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "需要安装 openpyxl: pip install openpyxl"
        ) from e

    p = Path(path)
    if not p.exists():
        return []

    # read_only=True 加速大文件
    wb = load_workbook(str(p), read_only=True, data_only=True)

    # 选择工作表
    if sheet is None:
        ws = wb.worksheets[0]
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]

    # 读取所有行（保留 None，便于检测表头）
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 决定列索引
    col_idx: int = 0
    header_consumed = False
    if column is not None:
        if isinstance(column, int):
            col_idx = column
        elif isinstance(column, str) and column.isascii() and column.isalpha():
            # 纯 ASCII 字母 → 列字母（A/B/C）；letter 直接定位，header 不跳过
            col_idx = _col_letter_to_index(column)
        else:
            # 字符串按 header 名匹配
            if not _is_header_row(rows[0]):
                raise ValueError(
                    f"指定了列名 {column!r} 但首行不是表头；"
                    "要么传列字母/列号，要么确保首行是表头"
                )
            headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            if column not in headers:
                raise ValueError(
                    f"列 {column!r} 不在表头中（{headers}）"
                )
            col_idx = headers.index(column)
            header_consumed = True

    # 若首行是表头，跳过（按列字母/列号时也跳过，避免取到 header 字符串）
    if _is_header_row(rows[0]):
        rows = rows[1:]

    # 提取
    seen: set[str] = set()
    out: list[str] = []
    for row in rows:
        if not row or col_idx >= len(row):
            continue
        cell = row[col_idx]
        if cell is None:
            continue
        s = str(cell).strip()
        if _is_comment(s):
            continue
        if s in seen:
            continue
        seen.add(s)
        out.append(s)

    wb.close()
    return out


def is_xlsx(path: str | Path) -> bool:
    return Path(path).suffix.lower() == ".xlsx"
