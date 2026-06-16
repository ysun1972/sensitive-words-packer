"""
敏感词脱敏工具 - 简易 GUI（tkinter + ttk）

特性：
- 单文件 / 目录输入
- 敏感词列表（.txt）/ Excel 词表（.xlsx）/ 规则文件（.json）
- 模式多选（精确 / 模糊 / 规则）
- 自定义通配符
- 实时日志（后台线程 + 队列避免 UI 冻结）
- 批量配置（JSON 任务清单）
- 5 个 tab：单次脱敏 / 批量任务 / 敏感词编辑 / 规则编辑 / Excel 词表编辑
- 启动时自动检查 + 安装依赖
- 全局异常捕获 → 写 stderr.log + 状态栏提示

设计：tkinter 是 Python 标准库，无需任何额外依赖。
"""
from __future__ import annotations

import json
import queue
import sys
import threading
import traceback
from pathlib import Path

# 允许直接 python src/gui.py
sys.path.insert(0, str(Path(__file__).parent))

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from core import SensitiveWordRedactor, load_words_file, load_rules_file
from excel_handler import load_excel_words
from file_handlers import supported_extensions
from batch import run_batch, BatchTask
from deps import check_deps, missing_deps, install_deps, InstallError


APP_TITLE = "敏感词脱敏工具 v0.2.2"
SUPPORTED = " ".join(supported_extensions())

# ====================== 依赖安装弹窗 ======================

class DepsInstallDialog(tk.Toplevel):
    """启动时如果缺依赖，弹这个窗口让用户确认安装"""

    def __init__(self, parent: tk.Tk, missing: list[str]):
        super().__init__(parent)
        self.title("需要安装依赖")
        self.geometry("640x440")
        self.transient(parent)
        self.grab_set()
        self.missing = missing
        self.user_confirmed = False
        self.install_thread: threading.Thread | None = None

        pad = {"padx": 8, "pady": 4}
        ttk.Label(
            self,
            text="缺少以下 Python 依赖，需要先安装：",
            font=("", 11, "bold"),
        ).pack(anchor=tk.W, **pad)
        ttk.Label(
            self,
            text="、".join(missing),
            foreground="#c00",
            font=("Menlo", 11),
        ).pack(anchor=tk.W, **pad)
        ttk.Label(
            self,
            text="将使用国内清华源加速下载（pip 默认会从 PyPI 拉）。",
            foreground="#666",
        ).pack(anchor=tk.W, **pad)
        ttk.Separator(self, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=8, pady=8)

        # 进度文本框
        self.txt = scrolledtext.ScrolledText(self, height=14, font=("Menlo", 9), wrap=tk.WORD)
        self.txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)
        self.txt.config(state=tk.DISABLED)
        self.queue: queue.Queue[str] = queue.Queue()

        # 按钮
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=8, pady=8)
        self.btn_install = ttk.Button(btn_frame, text="立即安装", command=self._on_install)
        self.btn_install.pack(side=tk.LEFT, padx=4)
        self.btn_cancel = ttk.Button(btn_frame, text="取消", command=self._on_cancel)
        self.btn_cancel.pack(side=tk.RIGHT, padx=4)
        self.var_status = tk.StringVar(value="等待用户操作")
        ttk.Label(btn_frame, textvariable=self.var_status, foreground="#666").pack(side=tk.LEFT, padx=8)

        self.after(100, self._poll)

    def _on_install(self):
        self.btn_install.config(state=tk.DISABLED)
        self.var_status.set("正在安装...")
        self.install_thread = threading.Thread(
            target=self._run_install, daemon=True
        )
        self.install_thread.start()

    def _run_install(self):
        try:
            install_deps(self.missing, progress_cb=self._enqueue)
            self.queue.put("__DONE__")
        except InstallError as e:
            self.queue.put(f"__ERROR__{e}")
        except Exception as e:
            self.queue.put(f"__ERROR__{e}")

    def _enqueue(self, line: str):
        self.queue.put(line)

    def _poll(self):
        try:
            while True:
                msg = self.queue.get_nowait()
                if msg.startswith("__DONE__"):
                    self.user_confirmed = True
                    self.var_status.set("安装成功")
                    self.after(500, self.destroy)
                    return
                if msg.startswith("__ERROR__"):
                    err = msg[len("__ERROR__"):]
                    self.var_status.set("安装失败")
                    self.btn_install.config(state=tk.NORMAL)
                    self._append_log(f"\n[错误] {err}\n")
                    messagebox.showerror("安装失败", f"{err}\n\n可手动执行：\npip install -i https://pypi.tuna.tsinghua.edu.cn/simple {' '.join(self.missing)}", parent=self)
                    return
                self._append_log(msg + "\n")
        except queue.Empty:
            pass
        self.after(100, self._poll)

    def _append_log(self, text: str):
        self.txt.config(state=tk.NORMAL)
        self.txt.insert(tk.END, text)
        self.txt.see(tk.END)
        self.txt.config(state=tk.DISABLED)

    def _on_cancel(self):
        self.user_confirmed = False
        self.destroy()


def run_deps_check_or_install(parent: tk.Tk) -> bool:
    """
    启动时调用：检查依赖，缺则弹窗让用户安装。
    返回 True=可继续启动 GUI；False=用户取消/失败。
    """
    report = check_deps()
    miss = missing_deps(report)
    if not miss:
        return True
    dlg = DepsInstallDialog(parent, miss)
    parent.wait_window(dlg)
    # 装完后重新检查一次（防止中途用户装了一半）
    if dlg.user_confirmed:
        return not missing_deps()
    return False


# ====================== 通用工具函数 ======================

def _read_text_file(path: str | Path) -> str:
    p = Path(path)
    if not p.exists():
        return ""
    for enc in ("utf-8", "utf-8-sig", "gbk", "gb18030"):
        try:
            return p.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return p.read_text(encoding="utf-8", errors="replace")


def _write_text_file(path: str | Path, content: str) -> None:
    Path(path).write_text(content, encoding="utf-8")


def _read_json_file(path: str | Path) -> dict | list:
    p = Path(path)
    if not p.exists():
        return {}
    for enc in ("utf-8", "utf-8-sig", "gbk"):
        try:
            return json.loads(p.read_text(encoding=enc))
        except UnicodeDecodeError:
            continue
    return json.loads(p.read_text(encoding="utf-8", errors="replace"))


# ====================== 内置示例 ======================

EXAMPLE_WORDS = """# 敏感词列表（每行一个，# 开头是注释）
张三
李四
王五
13800138000
admin@company.com
AcmeCorp
蓝海计划
"""

EXAMPLE_RULES = [
    {"name": "中国大陆手机号", "pattern": "(?<!\\d)1[3-9]\\d{9}(?!\\d)", "replacement": "[手机号]"},
    {"name": "电子邮箱", "pattern": "[\\w.+-]+@[\\w.-]+\\.[A-Za-z]{2,}", "replacement": "[邮箱]"},
    {"name": "18 位身份证号", "pattern": "(?<!\\d)\\d{17}[\\dXx](?!\\d)", "replacement": "[身份证]"},
    {"name": "银行卡号（16-19 位）", "pattern": "(?<!\\d)\\d{16,19}(?!\\d)", "replacement": "[银行卡]"},
    {"name": "IPv4 地址", "pattern": "(?<!\\d)(?:\\d{1,3}\\.){3}\\d{1,3}(?!\\d)", "replacement": "[IP]"},
]

EXAMPLE_EXCEL_DATA = [
    ["部门", "类型", "敏感词"],
    ["市场部", "客户", "AcmeCorp"],
    ["市场部", "客户", "Globex"],
    ["财务部", "账号", "6225880137460000"],
    ["财务部", "邮箱", "cfo@company.com"],
    ["研发部", "代号", "ProjectPhoenix"],
    ["研发部", "代号", "蓝海计划"],
]


# ====================== Treeview 可编辑支持 ======================

class EditableTreeview(ttk.Treeview):
    """支持双击单元格编辑的 Treeview"""

    def __init__(self, master=None, **kw):
        super().__init__(master, **kw)
        self._edit_entry: tk.Entry | None = None
        self.bind("<Double-1>", self._on_double_click)
        self.bind("<Button-1>", self._on_single_click, add="+")
        self.bind("<Return>", self._on_return)

    def _on_single_click(self, event):
        # 单击其他地方取消编辑
        if self._edit_entry and not self._is_over_entry(event):
            self._finish_edit()

    def _on_return(self, event):
        if self._edit_entry:
            self._finish_edit()
            return "break"
        return None

    def _is_over_entry(self, event) -> bool:
        if not self._edit_entry:
            return False
        try:
            x, y = self._edit_entry.winfo_rootx(), self._edit_entry.winfo_rooty()
            w, h = self._edit_entry.winfo_width(), self._edit_entry.winfo_height()
            return x <= event.x_root <= x + w and y <= event.y_root <= y + h
        except Exception:
            return False

    def _on_double_click(self, event):
        if self._edit_entry:
            self._finish_edit()
        region = self.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.identify_row(event.y)
        col_id = self.identify_column(event.x)
        if not row_id or not col_id:
            return
        col_index = int(col_id.replace("#", "")) - 1
        bbox = self.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        value = self.set(row_id, col_id)
        self._edit_entry = tk.Entry(self, borderwidth=1, relief=tk.SOLID)
        self._edit_entry.insert(0, value)
        self._edit_entry.select_range(0, tk.END)
        self._edit_entry.focus_set()
        self._edit_entry.place(x=x, y=y, width=w, height=h)
        self._edit_entry.bind("<Return>", lambda e: self._finish_edit())
        self._edit_entry.bind("<Escape>", lambda e: self._cancel_edit())
        self._edit_entry.bind("<FocusOut>", lambda e: self._finish_edit())
        self._edit_row = row_id
        self._edit_col = col_id

    def _finish_edit(self):
        if not self._edit_entry:
            return
        try:
            new_value = self._edit_entry.get()
            self.set(self._edit_row, self._edit_col, new_value)
        finally:
            self._cancel_edit()

    def _cancel_edit(self):
        if self._edit_entry:
            self._edit_entry.destroy()
            self._edit_entry = None


# ====================== 主窗口 ======================

class RedactorGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1000x750")
        self.root.minsize(900, 600)

        # 后台线程通信
        self.log_queue: queue.Queue[str] = queue.Queue()
        self.worker: threading.Thread | None = None

        # 状态变量
        self.var_input = tk.StringVar()
        self.var_output = tk.StringVar()
        self.var_words = tk.StringVar()
        self.var_excel = tk.StringVar()
        self.var_rules = tk.StringVar()
        self.var_batch = tk.StringVar()
        self.var_wildcard = tk.StringVar(value="***")
        self.mode_exact = tk.BooleanVar(value=True)
        self.mode_fuzzy = tk.BooleanVar(value=False)
        self.mode_rule = tk.BooleanVar(value=True)
        self.var_status = tk.StringVar(value="就绪")

        # 编辑器共享变量
        self.var_words_editor_path = tk.StringVar()
        self.var_rules_editor_path = tk.StringVar()
        self.var_excel_editor_path = tk.StringVar()
        self.var_excel_dirty = tk.BooleanVar(value=False)

        self._build_ui()
        self._poll_log_queue()

    # ---------------- UI 构建 ----------------

    def _build_ui(self):
        pad = {"padx": 6, "pady": 4}

        # 顶部说明
        top = ttk.Frame(self.root)
        top.pack(fill=tk.X, **pad)
        ttk.Label(
            top,
            text=f"支持格式：{SUPPORTED}",
            foreground="#666",
        ).pack(side=tk.LEFT)
        ttk.Label(
            top,
            text="  v0.2.2 — 自带依赖检查 + 内置编辑器",
            foreground="#999",
        ).pack(side=tk.LEFT, padx=10)

        # 模式 notebook
        nb = ttk.Notebook(self.root)
        nb.pack(fill=tk.BOTH, expand=True, **pad)

        self._build_single_tab(nb)
        self._build_batch_tab(nb)
        self._build_words_editor_tab(nb)
        self._build_rules_editor_tab(nb)
        self._build_excel_editor_tab(nb)

        # 底部：日志 + 状态栏
        log_frame = ttk.LabelFrame(self.root, text="运行日志")
        log_frame.pack(fill=tk.BOTH, expand=True, **pad)
        self.txt_log = scrolledtext.ScrolledText(
            log_frame, height=10, wrap=tk.WORD, font=("Menlo", 10)
        )
        self.txt_log.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.txt_log.config(state=tk.DISABLED)

        # 状态栏
        status_bar = ttk.Frame(self.root)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Label(status_bar, textvariable=self.var_status, anchor=tk.W).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=8, pady=2
        )

    def _build_single_tab(self, nb: ttk.Notebook):
        tab = ttk.Frame(nb)
        nb.add(tab, text="单次脱敏")
        pad = {"padx": 6, "pady": 4}

        row = 0
        ttk.Label(tab, text="输入路径（文件/目录）：").grid(row=row, column=0, sticky=tk.W, **pad)
        ttk.Entry(tab, textvariable=self.var_input).grid(row=row, column=1, sticky=tk.EW, **pad)
        ttk.Button(tab, text="选择...", command=self._pick_input).grid(row=row, column=2, **pad)

        row += 1
        ttk.Label(tab, text="输出目录：").grid(row=row, column=0, sticky=tk.W, **pad)
        ttk.Entry(tab, textvariable=self.var_output).grid(row=row, column=1, sticky=tk.EW, **pad)
        ttk.Button(tab, text="选择...", command=self._pick_output).grid(row=row, column=2, **pad)

        row += 1
        ttk.Separator(tab, orient=tk.HORIZONTAL).grid(
            row=row, column=0, columnspan=3, sticky=tk.EW, pady=8
        )

        row += 1
        ttk.Label(tab, text="敏感词列表（.txt）：").grid(row=row, column=0, sticky=tk.W, **pad)
        ttk.Entry(tab, textvariable=self.var_words).grid(row=row, column=1, sticky=tk.EW, **pad)
        ttk.Button(tab, text="选择...", command=self._pick_words).grid(row=row, column=2, **pad)

        row += 1
        ttk.Label(tab, text="Excel 词表（.xlsx）：").grid(row=row, column=0, sticky=tk.W, **pad)
        ttk.Entry(tab, textvariable=self.var_excel).grid(row=row, column=1, sticky=tk.EW, **pad)
        ttk.Button(tab, text="选择...", command=self._pick_excel).grid(row=row, column=2, **pad)

        row += 1
        ttk.Label(tab, text="规则文件（.json）：").grid(row=row, column=0, sticky=tk.W, **pad)
        ttk.Entry(tab, textvariable=self.var_rules).grid(row=row, column=1, sticky=tk.EW, **pad)
        ttk.Button(tab, text="选择...", command=self._pick_rules).grid(row=row, column=2, **pad)

        row += 1
        ttk.Separator(tab, orient=tk.HORIZONTAL).grid(
            row=row, column=0, columnspan=3, sticky=tk.EW, pady=8
        )

        row += 1
        ttk.Label(tab, text="脱敏模式：").grid(row=row, column=0, sticky=tk.W, **pad)
        mode_frame = ttk.Frame(tab)
        mode_frame.grid(row=row, column=1, columnspan=2, sticky=tk.W, **pad)
        ttk.Checkbutton(mode_frame, text="精确匹配", variable=self.mode_exact).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(mode_frame, text="模糊匹配", variable=self.mode_fuzzy).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(mode_frame, text="正则规则", variable=self.mode_rule).pack(side=tk.LEFT, padx=4)

        row += 1
        ttk.Label(tab, text="通配符：").grid(row=row, column=0, sticky=tk.W, **pad)
        ttk.Entry(tab, textvariable=self.var_wildcard, width=20).grid(
            row=row, column=1, sticky=tk.W, **pad
        )

        row += 1
        ttk.Separator(tab, orient=tk.HORIZONTAL).grid(
            row=row, column=0, columnspan=3, sticky=tk.EW, pady=8
        )

        row += 1
        btn_frame = ttk.Frame(tab)
        btn_frame.grid(row=row, column=0, columnspan=3, **pad)
        ttk.Button(btn_frame, text="开始脱敏", command=self._start_single).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="清空日志", command=self._clear_log).pack(side=tk.LEFT, padx=4)

        tab.columnconfigure(1, weight=1)

    def _build_batch_tab(self, nb: ttk.Notebook):
        tab = ttk.Frame(nb)
        nb.add(tab, text="批量任务")
        pad = {"padx": 6, "pady": 4}

        ttk.Label(
            tab,
            text="通过 JSON 任务清单一次跑多组脱敏任务，每任务可独立配置词表/规则/模式。\n"
                 "JSON 格式参见菜单「帮助」或 README.md。",
            foreground="#444",
            justify=tk.LEFT,
        ).grid(row=0, column=0, columnspan=3, sticky=tk.W, **pad)

        ttk.Label(tab, text="批量配置（.json）：").grid(row=1, column=0, sticky=tk.W, **pad)
        ttk.Entry(tab, textvariable=self.var_batch).grid(row=1, column=1, sticky=tk.EW, **pad)
        ttk.Button(tab, text="选择...", command=self._pick_batch).grid(row=1, column=2, **pad)

        ttk.Button(tab, text="生成示例配置", command=self._gen_batch_template).grid(
            row=2, column=1, sticky=tk.W, **pad
        )

        ttk.Separator(tab, orient=tk.HORIZONTAL).grid(
            row=3, column=0, columnspan=3, sticky=tk.EW, pady=8
        )

        btn_frame = ttk.Frame(tab)
        btn_frame.grid(row=4, column=0, columnspan=3, **pad)
        ttk.Button(btn_frame, text="开始批量", command=self._start_batch).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="清空日志", command=self._clear_log).pack(side=tk.LEFT, padx=4)

        tab.columnconfigure(1, weight=1)

    def _build_words_editor_tab(self, nb: ttk.Notebook):
        """敏感词编辑 tab"""
        tab = ttk.Frame(nb)
        nb.add(tab, text="📝 敏感词编辑")
        pad = {"padx": 6, "pady": 4}

        # 文件选择
        path_frame = ttk.Frame(tab)
        path_frame.pack(fill=tk.X, **pad)
        ttk.Label(path_frame, text="文件路径：").pack(side=tk.LEFT)
        ttk.Entry(path_frame, textvariable=self.var_words_editor_path).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=4
        )
        ttk.Button(path_frame, text="选择...", command=self._pick_words_editor).pack(side=tk.LEFT, padx=2)
        ttk.Button(path_frame, text="加载", command=self._load_words_editor).pack(side=tk.LEFT, padx=2)
        ttk.Button(path_frame, text="新建", command=self._new_words_editor).pack(side=tk.LEFT, padx=2)

        # 编辑区
        edit_frame = ttk.LabelFrame(tab, text="每行一个敏感词（# 开头是注释）")
        edit_frame.pack(fill=tk.BOTH, expand=True, **pad)
        self.txt_words = scrolledtext.ScrolledText(
            edit_frame, wrap=tk.NONE, font=("Menlo", 10)
        )
        self.txt_words.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # 按钮
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X, **pad)
        ttk.Button(btn_frame, text="保存", command=self._save_words_editor).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="另存为...", command=self._save_as_words_editor).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="清空", command=lambda: self._set_words_text("")).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="插入示例", command=lambda: self._set_words_text(EXAMPLE_WORDS)).pack(side=tk.LEFT, padx=4)

    def _build_rules_editor_tab(self, nb: ttk.Notebook):
        """规则编辑 tab"""
        tab = ttk.Frame(nb)
        nb.add(tab, text="🔧 规则编辑")
        pad = {"padx": 6, "pady": 4}

        # 文件选择
        path_frame = ttk.Frame(tab)
        path_frame.pack(fill=tk.X, **pad)
        ttk.Label(path_frame, text="文件路径：").pack(side=tk.LEFT)
        ttk.Entry(path_frame, textvariable=self.var_rules_editor_path).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=4
        )
        ttk.Button(path_frame, text="选择...", command=self._pick_rules_editor).pack(side=tk.LEFT, padx=2)
        ttk.Button(path_frame, text="加载", command=self._load_rules_editor).pack(side=tk.LEFT, padx=2)
        ttk.Button(path_frame, text="新建", command=self._new_rules_editor).pack(side=tk.LEFT, padx=2)

        # 表格
        table_frame = ttk.LabelFrame(tab, text="正则规则（双击单元格编辑）")
        table_frame.pack(fill=tk.BOTH, expand=True, **pad)
        self.tree_rules = EditableTreeview(
            table_frame,
            columns=("name", "pattern", "replacement"),
            show="headings",
            height=12,
        )
        self.tree_rules.heading("name", text="名称")
        self.tree_rules.heading("pattern", text="正则表达式")
        self.tree_rules.heading("replacement", text="替换为")
        self.tree_rules.column("name", width=180, anchor=tk.W)
        self.tree_rules.column("pattern", width=380, anchor=tk.W)
        self.tree_rules.column("replacement", width=200, anchor=tk.W)
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree_rules.yview)
        self.tree_rules.configure(yscrollcommand=vsb.set)
        self.tree_rules.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=4)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # 按钮
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X, **pad)
        ttk.Button(btn_frame, text="添加行", command=self._add_rule_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="删除选中", command=self._del_rule_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="上移", command=lambda: self._move_rule_row(-1)).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="下移", command=lambda: self._move_rule_row(1)).pack(side=tk.LEFT, padx=4)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="保存", command=self._save_rules_editor).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="另存为...", command=self._save_as_rules_editor).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="加载示例", command=self._load_example_rules).pack(side=tk.LEFT, padx=4)

    def _build_excel_editor_tab(self, nb: ttk.Notebook):
        """Excel 词表编辑 tab"""
        tab = ttk.Frame(nb)
        nb.add(tab, text="📊 Excel 词表编辑")
        pad = {"padx": 6, "pady": 4}

        # 文件选择
        path_frame = ttk.Frame(tab)
        path_frame.pack(fill=tk.X, **pad)
        ttk.Label(path_frame, text="文件路径：").pack(side=tk.LEFT)
        ttk.Entry(path_frame, textvariable=self.var_excel_editor_path).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=4
        )
        ttk.Button(path_frame, text="选择...", command=self._pick_excel_editor).pack(side=tk.LEFT, padx=2)
        ttk.Button(path_frame, text="加载", command=self._load_excel_editor).pack(side=tk.LEFT, padx=2)
        ttk.Button(path_frame, text="新建", command=self._new_excel_editor).pack(side=tk.LEFT, padx=2)

        # 表格（动态列）
        table_frame = ttk.LabelFrame(tab, text="词表（双击单元格编辑）")
        table_frame.pack(fill=tk.BOTH, expand=True, **pad)
        self.tree_excel = EditableTreeview(table_frame, show="headings", height=12)
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree_excel.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree_excel.xview)
        self.tree_excel.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_excel.grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self._excel_columns: list[str] = []

        # 按钮
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X, **pad)
        ttk.Button(btn_frame, text="添加行", command=self._add_excel_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="删除选中", command=self._del_excel_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="添加列", command=self._add_excel_column).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="删除列", command=self._del_excel_column).pack(side=tk.LEFT, padx=4)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="保存", command=self._save_excel_editor).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="另存为...", command=self._save_as_excel_editor).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="加载示例", command=self._load_example_excel).pack(side=tk.LEFT, padx=4)

    # ---------------- 文件选择 ----------------

    def _pick_input(self):
        path = filedialog.askopenfilename(title="选择输入文件（取消选目录）")
        if not path:
            path = filedialog.askdirectory(title="或选择输入目录")
        if path:
            self.var_input.set(path)

    def _pick_output(self):
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.var_output.set(path)

    def _pick_words(self):
        path = filedialog.askopenfilename(
            title="选择敏感词列表", filetypes=[("文本", "*.txt"), ("所有", "*.*")]
        )
        if path:
            self.var_words.set(path)

    def _pick_excel(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 词表", filetypes=[("Excel", "*.xlsx"), ("所有", "*.*")]
        )
        if path:
            self.var_excel.set(path)

    def _pick_rules(self):
        path = filedialog.askopenfilename(
            title="选择规则文件", filetypes=[("JSON", "*.json"), ("所有", "*.*")]
        )
        if path:
            self.var_rules.set(path)

    def _pick_batch(self):
        path = filedialog.askopenfilename(
            title="选择批量配置", filetypes=[("JSON", "*.json"), ("所有", "*.*")]
        )
        if path:
            self.var_batch.set(path)

    def _gen_batch_template(self):
        out = filedialog.asksaveasfilename(
            title="保存批量配置示例",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile="batch_example.json",
        )
        if not out:
            return
        template = {
            "_comment": "sensitive-words-packer 批量任务示例；删除 _comment 行",
            "tasks": [
                {
                    "name": "示例任务 1 - 文本脱敏",
                    "input": "./input/marketing",
                    "output": "./output/marketing",
                    "words": "./config/words.txt",
                    "rules": "./config/rules.json",
                    "mode": "exact,rule",
                    "wildcard": "***",
                    "log": "marketing.log",
                },
                {
                    "name": "示例任务 2 - Excel 词表",
                    "input": "./input/finance",
                    "output": "./output/finance",
                    "excel": "./config/words_demo.xlsx",
                    "excel_sheet": "敏感词",
                    "excel_column": "敏感词",
                    "rules": "./config/rules_finance.json",
                    "mode": "exact,fuzzy,rule",
                    "wildcard": "[REDACTED]",
                },
            ],
        }
        Path(out).write_text(
            json.dumps(template, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        self.var_batch.set(out)
        self._log(f"已生成批量配置示例: {out}")

    # ---------------- 敏感词编辑器 ----------------

    def _pick_words_editor(self):
        path = filedialog.askopenfilename(
            title="选择敏感词文件", filetypes=[("文本", "*.txt"), ("所有", "*.*")]
        )
        if path:
            self.var_words_editor_path.set(path)
            self._load_words_editor()

    def _new_words_editor(self):
        self.var_words_editor_path.set("")
        self._set_words_text(EXAMPLE_WORDS)

    def _load_words_editor(self):
        path = self.var_words_editor_path.get().strip()
        if not path:
            messagebox.showwarning("提示", "请先选择文件路径")
            return
        try:
            content = _read_text_file(path)
            self._set_words_text(content)
            self._log(f"[敏感词] 已加载 {path}（{len(content.splitlines())} 行）")
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _set_words_text(self, content: str):
        self.txt_words.config(state=tk.NORMAL)
        self.txt_words.delete("1.0", tk.END)
        self.txt_words.insert("1.0", content)
        self.txt_words.config(state=tk.NORMAL)  # 保持可编辑

    def _get_words_text(self) -> str:
        return self.txt_words.get("1.0", "end-1c")

    def _save_words_editor(self):
        path = self.var_words_editor_path.get().strip()
        if not path:
            self._save_as_words_editor()
            return
        try:
            _write_text_file(path, self._get_words_text())
            self._log(f"[敏感词] 已保存 {path}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def _save_as_words_editor(self):
        path = filedialog.asksaveasfilename(
            title="另存为敏感词文件",
            defaultextension=".txt",
            filetypes=[("文本", "*.txt"), ("所有", "*.*")],
            initialfile="words.txt",
        )
        if not path:
            return
        try:
            _write_text_file(path, self._get_words_text())
            self.var_words_editor_path.set(path)
            self._log(f"[敏感词] 已保存 {path}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    # ---------------- 规则编辑器 ----------------

    def _pick_rules_editor(self):
        path = filedialog.askopenfilename(
            title="选择规则文件", filetypes=[("JSON", "*.json"), ("所有", "*.*")]
        )
        if path:
            self.var_rules_editor_path.set(path)
            self._load_rules_editor()

    def _new_rules_editor(self):
        self.var_rules_editor_path.set("")
        for item in self.tree_rules.get_children():
            self.tree_rules.delete(item)
        self._load_example_rules()

    def _load_rules_editor(self):
        path = self.var_rules_editor_path.get().strip()
        if not path:
            messagebox.showwarning("提示", "请先选择文件路径")
            return
        try:
            data = _read_json_file(path)
            items = data["rules"] if isinstance(data, dict) and "rules" in data else data
            for item in self.tree_rules.get_children():
                self.tree_rules.delete(item)
            for r in items:
                self.tree_rules.insert("", tk.END, values=(
                    r.get("name", ""),
                    r.get("pattern", ""),
                    r.get("replacement", "***"),
                ))
            self._log(f"[规则] 已加载 {path}（{len(items)} 条）")
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _add_rule_row(self):
        self.tree_rules.insert("", tk.END, values=("新规则", "", "***"))

    def _del_rule_row(self):
        for item in self.tree_rules.selection():
            self.tree_rules.delete(item)

    def _move_rule_row(self, delta: int):
        for item in self.tree_rules.selection():
            idx = self.tree_rules.index(item)
            new_idx = max(0, idx + delta)
            self.tree_rules.move(item, "", new_idx)

    def _get_rules_data(self) -> list[dict]:
        rules = []
        for item in self.tree_rules.get_children():
            values = self.tree_rules.item(item, "values")
            if values and (values[0] or values[1]):
                rules.append({
                    "name": values[0],
                    "pattern": values[1],
                    "replacement": values[2] if len(values) > 2 else "***",
                })
        return rules

    def _save_rules_editor(self):
        path = self.var_rules_editor_path.get().strip()
        if not path:
            self._save_as_rules_editor()
            return
        try:
            data = {"rules": self._get_rules_data()}
            _write_text_file(path, json.dumps(data, ensure_ascii=False, indent=2))
            self._log(f"[规则] 已保存 {path}（{len(data['rules'])} 条）")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def _save_as_rules_editor(self):
        path = filedialog.asksaveasfilename(
            title="另存为规则文件",
            defaultextension=".json",
            filetypes=[("JSON", "*.json"), ("所有", "*.*")],
            initialfile="rules.json",
        )
        if not path:
            return
        try:
            data = {"rules": self._get_rules_data()}
            _write_text_file(path, json.dumps(data, ensure_ascii=False, indent=2))
            self.var_rules_editor_path.set(path)
            self._log(f"[规则] 已保存 {path}（{len(data['rules'])} 条）")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def _load_example_rules(self):
        for item in self.tree_rules.get_children():
            self.tree_rules.delete(item)
        for r in EXAMPLE_RULES:
            self.tree_rules.insert("", tk.END, values=(r["name"], r["pattern"], r["replacement"]))

    # ---------------- Excel 词表编辑器 ----------------

    def _rebuild_excel_tree(self, columns: list[str]):
        """重置 Treeview 的列定义"""
        self._excel_columns = list(columns)
        self.tree_excel.configure(columns=columns if columns else ())
        for c in columns:
            self.tree_excel.heading(c, text=c)
            self.tree_excel.column(c, width=140, anchor=tk.W)

    def _pick_excel_editor(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 词表", filetypes=[("Excel", "*.xlsx"), ("所有", "*.*")]
        )
        if path:
            self.var_excel_editor_path.set(path)
            self._load_excel_editor()

    def _new_excel_editor(self):
        self.var_excel_editor_path.set("")
        self._load_example_excel()

    def _load_excel_editor(self):
        path = self.var_excel_editor_path.get().strip()
        if not path:
            messagebox.showwarning("提示", "请先选择文件路径")
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                self._rebuild_excel_tree(["列1"])
                return
            # 首行作为 header
            header = [str(c).strip() if c is not None else f"列{i+1}" for i, c in enumerate(rows[0])]
            self._rebuild_excel_tree(header)
            for item in self.tree_excel.get_children():
                self.tree_excel.delete(item)
            for row in rows[1:]:
                values = [str(c) if c is not None else "" for c in row]
                # 对齐列数
                while len(values) < len(header):
                    values.append("")
                self.tree_excel.insert("", tk.END, values=values[:len(header)])
            self._log(f"[Excel] 已加载 {path}（{len(rows)-1} 行 × {len(header)} 列）")
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _load_example_excel(self):
        self._rebuild_excel_tree(EXAMPLE_EXCEL_DATA[0])
        for item in self.tree_excel.get_children():
            self.tree_excel.delete(item)
        for row in EXAMPLE_EXCEL_DATA[1:]:
            self.tree_excel.insert("", tk.END, values=list(row))

    def _add_excel_row(self):
        n = len(self._excel_columns)
        if n == 0:
            self._rebuild_excel_tree(["列1"])
            n = 1
        self.tree_excel.insert("", tk.END, values=[""] * n)

    def _del_excel_row(self):
        for item in self.tree_excel.selection():
            self.tree_excel.delete(item)

    def _add_excel_column(self):
        from tkinter import simpledialog
        if not self._excel_columns:
            self._rebuild_excel_tree(["新列"])
            return
        name = simpledialog.askstring("新列名", "请输入新列名：", parent=self.tree_excel)
        if not name:
            return
        self._excel_columns.append(name)
        self.tree_excel.configure(columns=self._excel_columns)
        self.tree_excel.heading(name, text=name)
        self.tree_excel.column(name, width=140, anchor=tk.W)
        # 给已有行填充空值
        for item in self.tree_excel.get_children():
            values = list(self.tree_excel.item(item, "values"))
            values.append("")
            self.tree_excel.item(item, values=values)

    def _del_excel_column(self):
        if not self._excel_columns:
            return
        from tkinter import simpledialog
        name = simpledialog.askstring(
            "删除列",
            f"请输入要删除的列名：\n{', '.join(self._excel_columns)}",
            parent=self.tree_excel,
        )
        if not name or name not in self._excel_columns:
            return
        idx = self._excel_columns.index(name)
        self._excel_columns.pop(idx)
        self.tree_excel.configure(columns=self._excel_columns)
        for item in self.tree_excel.get_children():
            values = list(self.tree_excel.item(item, "values"))
            if idx < len(values):
                values.pop(idx)
            self.tree_excel.item(item, values=values)
        if self._excel_columns:
            for c in self._excel_columns:
                self.tree_excel.heading(c, text=c)
                self.tree_excel.column(c, width=140, anchor=tk.W)

    def _get_excel_data(self) -> tuple[list[str], list[list]]:
        """返回 (header, rows)"""
        rows = []
        for item in self.tree_excel.get_children():
            values = list(self.tree_excel.item(item, "values"))
            # 对齐列数
            while len(values) < len(self._excel_columns):
                values.append("")
            rows.append(values[:len(self._excel_columns)])
        return list(self._excel_columns), rows

    def _save_excel_editor(self):
        path = self.var_excel_editor_path.get().strip()
        if not path:
            self._save_as_excel_editor()
            return
        try:
            self._write_excel(path)
            self._log(f"[Excel] 已保存 {path}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def _save_as_excel_editor(self):
        path = filedialog.asksaveasfilename(
            title="另存为 Excel 词表",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("所有", "*.*")],
            initialfile="words.xlsx",
        )
        if not path:
            return
        try:
            self._write_excel(path)
            self.var_excel_editor_path.set(path)
            self._log(f"[Excel] 已保存 {path}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def _write_excel(self, path: str):
        from openpyxl import Workbook
        header, rows = self._get_excel_data()
        wb = Workbook()
        ws = wb.active
        ws.title = "敏感词"
        if header:
            ws.append(header)
        for r in rows:
            ws.append(r)
        wb.save(path)

    # ---------------- 后台任务 ----------------

    def _start_single(self):
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("提示", "有任务正在运行，请等待完成")
            return

        input_path = self.var_input.get().strip()
        output_dir = self.var_output.get().strip()
        if not input_path or not output_dir:
            messagebox.showerror("错误", "请填写输入路径和输出目录")
            return

        # 收集模式
        modes: list[str] = []
        if self.mode_exact.get():
            modes.append("word-exact")
        if self.mode_fuzzy.get():
            modes.append("word-fuzzy")
        if self.mode_rule.get():
            modes.append("rule")
        if not modes:
            messagebox.showerror("错误", "请至少启用一个脱敏模式")
            return

        # 加载配置
        try:
            words: list[str] = []
            if self.var_words.get().strip():
                words.extend(load_words_file(self.var_words.get().strip()))
            if self.var_excel.get().strip():
                words.extend(load_excel_words(self.var_excel.get().strip()))
            rules = load_rules_file(self.var_rules.get().strip()) if self.var_rules.get().strip() else []
        except Exception as e:
            messagebox.showerror("加载配置失败", str(e))
            return

        if not words and not rules:
            messagebox.showerror("错误", "请至少提供敏感词列表、Excel 词表或规则文件之一")
            return

        wildcard = self.var_wildcard.get() or "***"
        redactor = SensitiveWordRedactor(words=words, rules=rules, wildcard=wildcard)
        self._log(f"[配置] {len(words)} 词 | {len(rules)} 规则 | 模式 {modes} | 通配符 {wildcard!r}")

        # 后台线程跑
        self.var_status.set("处理中...")
        self.worker = threading.Thread(
            target=self._run_single_worker,
            args=(Path(input_path), Path(output_dir), redactor, modes),
            daemon=True,
        )
        self.worker.start()

    def _run_single_worker(self, input_path: Path, output_dir: Path,
                           redactor: SensitiveWordRedactor, modes: list[str]):
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            files = [input_path] if input_path.is_file() else \
                [p for p in input_path.rglob("*") if p.is_file()]
            self.log_queue.put(f"[扫描] {len(files)} 个文件")
            total = 0
            log_lines: list[str] = []
            for f in files:
                self.log_queue.put(f"  → {f.name}")
                from file_handlers import get_handler, copy_file
                handler = get_handler(f)
                if handler is None:
                    outs = copy_file(f, output_dir)
                    self.log_queue.put(f"    [⤳ 复制] {f.name}（不支持格式）")
                    continue
                try:
                    text = handler.read(f)
                    result = redactor.redact(text, source_file=f.name, modes=modes)
                    outs = handler.write(output_dir / f.name, result.redacted_text)
                    n = len(result.matches)
                    total += n
                    extra = ""
                    if len(outs) > 1:
                        extra = " [+ " + ", ".join(o.name for o in outs[1:]) + "]"
                    self.log_queue.put(f"    [✓ 完成] {f.name}  {n} 处匹配{extra}")
                    for m in result.matches:
                        log_lines.append(
                            f"[{m.mode}{':' + m.rule_name if m.rule_name else ''}] "
                            f"{m.file}:{m.line}  '{m.original}' -> '{m.replacement}'"
                        )
                except Exception as e:
                    self.log_queue.put(f"    [✗ 失败] {f.name}  {e}")
            log_path = output_dir / "run.log"
            log_path.write_text(
                f"# sensitive-words-packer 审计日志\n"
                f"# 输入: {input_path}\n"
                f"# 模式: {modes}\n"
                f"# 文件数: {len(files)} | 总匹配: {total}\n"
                f"\n" + "\n".join(log_lines),
                encoding="utf-8",
            )
            self.log_queue.put(f"[完成] 输出目录: {output_dir}")
            self.log_queue.put(f"[完成] 总匹配: {total} 处 | 审计日志: {log_path}")
            self.log_queue.put("__STATUS__done__")
        except Exception as e:
            self.log_queue.put(f"[错误] {e}")
            self.log_queue.put(traceback.format_exc())
            self.log_queue.put("__STATUS__done__")

    def _start_batch(self):
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("提示", "有任务正在运行，请等待完成")
            return
        cfg = self.var_batch.get().strip()
        if not cfg:
            messagebox.showerror("错误", "请选择批量配置文件")
            return

        self.var_status.set("批量处理中...")
        self.worker = threading.Thread(
            target=self._run_batch_worker, args=(cfg,), daemon=True
        )
        self.worker.start()

    def _run_batch_worker(self, cfg_path: str):
        try:
            def progress(idx, total, name):
                self.log_queue.put(f"[批量] ({idx}/{total}) {name}")
            results = run_batch(cfg_path, progress_cb=progress)
            ok = sum(1 for r in results if r.status == "ok")
            err = sum(1 for r in results if r.status != "ok")
            for r in results:
                if r.status == "ok":
                    self.log_queue.put(
                        f"  [✓ {r.task.name}] {r.files} 文件 / {r.matches} 匹配"
                    )
                else:
                    self.log_queue.put(f"  [✗ {r.task.name}] {r.error}")
            self.log_queue.put(f"[批量完成] 成功 {ok} | 失败 {err}")
            self.log_queue.put("__STATUS__done__")
        except Exception as e:
            self.log_queue.put(f"[批量错误] {e}")
            self.log_queue.put(traceback.format_exc())
            self.log_queue.put("__STATUS__done__")

    # ---------------- 日志 + 状态 ----------------

    def _log(self, msg: str):
        self.txt_log.config(state=tk.NORMAL)
        self.txt_log.insert(tk.END, msg + "\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state=tk.DISABLED)

    def _clear_log(self):
        self.txt_log.config(state=tk.NORMAL)
        self.txt_log.delete("1.0", tk.END)
        self.txt_log.config(state=tk.DISABLED)

    def _poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if msg == "__STATUS__done__":
                    self.var_status.set("就绪")
                else:
                    self._log(msg)
        except queue.Empty:
            pass
        self.root.after(100, self._poll_log_queue)


# ====================== 入口 ======================

def main():
    """GUI 入口；异常兜底写 stderr.log 便于排错"""
    try:
        root = tk.Tk()
        try:
            style = ttk.Style()
            if "vista" in style.theme_names():
                style.theme_use("vista")
            elif "clam" in style.theme_names():
                style.theme_use("clam")
        except Exception:
            pass

        # 启动时检查依赖
        if not run_deps_check_or_install(root):
            root.destroy()
            sys.exit(1)

        # 隐藏 root 直到 RedactorGUI 创建完成（避免窗口闪烁）
        RedactorGUI(root)
        root.mainloop()
    except Exception as e:
        err = traceback.format_exc()
        try:
            Path("stderr.log").write_text(err, encoding="utf-8")
        except Exception:
            pass
        try:
            if sys.stderr is not None:
                print(err, file=sys.stderr)
        except Exception:
            pass
        try:
            messagebox.showerror("启动失败", f"{e}\n\n详见 stderr.log")
        except Exception:
            pass
        sys.exit(1)


if __name__ == "__main__":
    main()
